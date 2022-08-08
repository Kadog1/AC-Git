# -*- coding: utf-8 -*-
"""
Created on Mon Feb 21 12:04:37 2022

@author: Paul Ferdinand Popp, Assistant, CAD
"""

from datetime import datetime
import fnmatch
import logging
import openpyxl
import os
import pandas as pd
import pyodbc


isProductRun = False
if isProductRun == True:
    tTEST = ""
    Testumgebung = ""
else:
    tTEST = "_TEST"
    Testumgebung = " Testumgebung"
    
formatter = logging.Formatter(fmt="%(asctime)s - %(levelname)s: %(message)s", datefmt="%Y-%m-%d %H:%M:%S")    
def setup_logger(name, log_file, level=logging.INFO):
    handler = logging.FileHandler(log_file)        
    handler.setFormatter(formatter)
    logger = logging.getLogger(name)
    logger.setLevel(level)
    logger.addHandler(handler)
    return logger

# Find all files that match the pattern in path
def find(pattern, path):
    try:
        result = []
        for root, dirs, files in os.walk(path):
            for name in files:
                if fnmatch.fnmatch(name, pattern):
                    result.append(os.path.join(root, name))
        return result
    except Exception as e:
        logging.error('Error in find: '+ str(e))
        raise

# Returns file name without prohibited characters
def getFilename(*args):
    try:
        pattern = ['Ä', 'ä', 'Ö', 'ö', 'Ü', 'ü', '~', '“', '#', '%', '&', '*', ':', '<', '>', '?', '/', '\\', '{', '|', '}']
        replace = ['Ae', 'ae', 'Oe', 'oe', 'Ue', 'ue', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
        fileName = "_".join(list(args))
        for argPattern, argReplace in zip(pattern, replace):
            fileName = fileName.replace(argPattern, argReplace)
        return fileName
    except Exception as e:
        logging.error('Error in getFilename: '+ str(e))
        raise

# Returns directory path with prohibited characters. If path doesnt exist, create path.
def getSaveDir(compName):
    try:
        savePath = r'\\Defrnappfl101.ey.net\101fra00010\T\TCC_SB\Z_Archive\Adressabgleich' + Testumgebung + '\F Screenshots'
        pattern = ['Ä', 'ä', 'Ö', 'ö', 'Ü', 'ü', '~', '“', '#', '%', '&', '*', ':', '<', '>', '?', '/', '\\', '{', '|', '}']
        replace = ['Ae', 'ae', 'Oe', 'oe', 'Ue', 'ue', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
        for argPattern, argReplace in zip(pattern, replace):
            compName = compName.replace(argPattern, argReplace)
        savePath = savePath + '\\' + str(compName[0]).upper() + '\\' + str(compName) + '\\'
        if os.path.isdir(savePath) == False:
            os.makedirs(savePath)
        return savePath
    except Exception as e:
        logging.error('Error in getSaveDir: '+ str(e))
        raise
        

# If address found at register website, write back to sql database and create new entries
def createSQLAddressEntry(compName, street, zipCode, location, country, orderNo, idxAddress, saveDir, fileName, timeStamp, register):
    try:
        connSQLServer = pyodbc.connect('Driver={ODBC Driver 11 for SQL Server};'
                              'Server=DERUSCMPDWASQ01.ey.net\INST02;'
                              'Database=CAD;'
                              'Trusted_Connection=yes;')
        cursor = connSQLServer.cursor()
        timeStamp = timeStamp[0:4] + '-' + timeStamp[4:6] + '-' + timeStamp[6:8] + ' ' + timeStamp[9:11] + ':' + timeStamp[11:13]
        
        # Get keyAddress
        sqlFindKeyAddress = "SELECT TOP (1) keyAddress FROM [CAD].[dbo].[tAC_Addresses{5}] WHERE Company = '{0}' \
                            AND Strasse_Hausnummer_Postfach = '{1}' AND Postleitzahl = '{2}' AND Stadt = '{3}' \
                            AND Land = '{4}'".format(compName, street, zipCode, location, 'DE', tTEST)
        cursor.execute(sqlFindKeyAddress)
        dataRecordSet = []
        for row in cursor.fetchall():
            row_to_list = [elem for elem in row]
            dataRecordSet.append(row_to_list)
        dfRecordSet = pd.DataFrame(dataRecordSet)
        if len(dfRecordSet) == 0:
            sqlMaxKeyAddress = "SELECT TOP (1) keyAddress FROM [CAD].[dbo].[tAC_Addresses{0}] ORDER BY LEN(keyAddress) DESC, keyAddress DESC".format(tTEST)
            cursor.execute(sqlMaxKeyAddress)
            dataRecordSet = []
            for row in cursor.fetchall():
                row_to_list = [elem for elem in row]
                dataRecordSet.append(row_to_list)
            dfRecordSet = pd.DataFrame(dataRecordSet)
            if len(dfRecordSet) == 0: 
                keyAddress = 'A1'
            else:
                maxKeyAddress = dfRecordSet.iloc[0,0]
                keyAddress = 'A' + str(int(maxKeyAddress[1:]) + 1)
        else:
            keyAddress = dfRecordSet.iloc[0,0]
            
        sqlInsertAddress = "INSERT INTO [CAD].[dbo].[tAC_Addresses{6}] (Company, Strasse_Hausnummer_Postfach, Postleitzahl, Stadt, Land, keyAddress) \
                            SELECT TOP (1) '{0}', '{1}', '{2}', '{3}', '{4}', '{5}' FROM [CAD].[dbo].[tAC_Addresses{6}] \
                            WHERE NOT EXISTS (SELECT * FROM [CAD].[dbo].[tAC_Addresses{6}] WHERE \
                            Company = '{0}' AND Strasse_Hausnummer_Postfach = '{1}' AND Postleitzahl = '{2}' AND Stadt = '{3}' \
                            AND Land = '{4}')".format(compName, street, zipCode, location, country, keyAddress, tTEST)                   
        sqlInsertScreenshot = "INSERT INTO [CAD].[dbo].[tAC_Screenshots{6}] (Company, keyAddress, tsScreenshotCreated, locationPath, pngFileName, register)\
                       VALUES ('{0}', '{4}', '{1}', '{2}', '{3}', '{5}')".format(compName, timeStamp, saveDir, fileName, keyAddress, register, tTEST)                  
        sqlInsertProdScreenshot = "INSERT INTO [CAD].[dbo].[tAC_ProdScreenshots{4}] (Company, OrderNo, idxAddress, tsScreenshotCreated)\
                       VALUES ('{0}', '{1}', {2}, '{3}')".format(compName, orderNo, idxAddress, timeStamp, tTEST)
                       
        cursor.execute(sqlInsertAddress)
        cursor.execute(sqlInsertScreenshot)
        cursor.execute(sqlInsertProdScreenshot)
        connSQLServer.commit()
    except Exception as e:
        logging.error('Error in createSQLAddressEntry: '+ str(e))
        raise
    
def getRangeBySheetName(sheetName):
    try:
        if sheetName == 'Adresscheck' or sheetName == 'Rechts-_Steuerberater' or sheetName == 'Address check' or sheetName == 'Law_Tax Advisor' or sheetName == 'Legal_Tax Advisors':
            strRange = 'C:L' # Name der Firma, Adresszusatz, Ansprechpartner - Vorname, Ansprechpartner - Nachname, Straße + Hausnummer / Postfach, Postleitzahl, Stadt, Land, E-Mail					
        elif sheetName == 'Debitor_Kreditor_Sonst':
            strRange = 'D:M'
        elif sheetName == 'Bank':
            strRange = 'D:M'
        else:
            strRange = 'D:M'
        return strRange
    except Exception as e:
        logging.error('Error in getRangeBySheetName: '+ str(e))
        raise

# Load list of addresses from Documentation Workbook for a specific orderNo
def getDFWorkbook(orderNo):
    try:
        df = pd.DataFrame()
        timeStamp = datetime.now().replace(second=0, microsecond=0)
        pathToCADAbgleich = "\\\Defrnappfl101.ey.net\\101fra00010\\T\\TCC_SB\\Z_Archive\\eConfirmations\\Datenbank\\C Workplace\\" + orderNo + r"\2. CAD_Abgleich"
        pathToWorkbook =  find('*1_CAD-Adressabgleich Adressenabfrage Mandant*', pathToCADAbgleich)[0]
        if pathToWorkbook[-4:] == 'xlsm':
            wbInputSheet = openpyxl.load_workbook(pathToWorkbook, keep_vba=True)
        else:
            wbInputSheet = openpyxl.load_workbook(pathToWorkbook, keep_vba=False)
        wsBasicInfo = wbInputSheet['basic_info']
        wsBasicInfo['E3'] = timeStamp
        wbInputSheet.save(pathToWorkbook)
        sheetNames = pd.ExcelFile(pathToWorkbook).sheet_names
        sheetNames = [x for x in sheetNames if x not in ['Summary', 'ISO', 'basic_info', 'Inhalte']]
        for sheetName in sheetNames:
            strRange = getRangeBySheetName(sheetName)
            dfAppend = pd.read_excel(pathToWorkbook, sheetName, skiprows = 10, usecols = strRange, header=None)
            dfAppend.columns = range(dfAppend.columns.size)
            firstRow = 6
            for i in range(dfAppend.columns.size):
                try:
                    firstRow = dfAppend.loc[(dfAppend[i].apply(lambda x: str(x).upper() == 'ADDITIONAL ADDRESS INFORMATION' or str(x).upper() == 'ADRESSZUSATZ'))].reset_index()['index'][0] + 2
                    break
                except Exception:
                    pass
            dfAppend = dfAppend.iloc[firstRow:, :]
            df = df.append(dfAppend, ignore_index=True)
        if len(df) == 0: 
            return None
        df.columns = ['Confirmation', 'Company', 'Adresszusatz', 'Surname', 'Lastname', 'Street', 'Zipcode', 'Location', 'Country', 'Email']
        df.dropna(axis = 0, how = 'all', inplace = True)
        df = df.drop_duplicates(subset=['Company', 'Surname', 'Lastname', 'Street', 'Zipcode', 'Location', 'Country', 'Email'])
        df = df.reset_index(drop=True)
        df = df.reset_index(drop=False)
        #df = df.loc[df.loc[:, 'Country'].str.upper() == 'DE', :]
        
        df = df.loc[df['Company'].apply(lambda x: type(x) is str), :] # enforce datatype #KA datatypeFix for addresses with IDs 202206011
        df.loc[:, 'Location'] = df.loc[:, 'Location'].fillna('location not specified')   
        df = df.loc[df['Country'].apply(lambda x: type(x) is str), :]
        df['OrderNo'] = orderNo
        df['timeStamp'] = datetime.strftime(timeStamp, '%Y%m%d_%H%M')
        return df
    
    except Exception as e:
        logging.error('Error in getDFWorkbook: '+ str(e))
        raise
    